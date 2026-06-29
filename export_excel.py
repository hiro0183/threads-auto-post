"""
【コンサル垢 専用】
Threads投稿内容をExcelシートに出力するスクリプト

※ ラポール整体院（@rapport.sango）側とは別物。混同しないこと。

使い方:
  python export_excel.py              # 明日から30日分をデスクトップに出力
  python export_excel.py 2026-04-01  # 指定日から30日分
  python export_excel.py 2026-04-01 60  # 指定日から60日分
"""

import sys
import json
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# post_runnerからSLOT_PLANをインポート
sys.path.insert(0, str(Path(__file__).parent))
from post_runner import SLOT_PLAN, POST_SCHEDULE

BASE_DIR = Path(__file__).parent
POSTS_DIR = BASE_DIR / "posts"

# デスクトップパス（OneDrive優先）
DESKTOP = Path.home() / "OneDrive" / "Desktop"
if not DESKTOP.exists():
    DESKTOP = Path.home() / "Desktop"


# ── スタイル定義 ────────────────────────────────────────

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# 日付ごとの背景色（交互）
DATE_COLORS = ["F0F4FF", "FFF8F0"]

HEADER_FILL  = make_fill("2C3E50")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
CTA_FILL     = make_fill("FFF3CD")
NEW_SLOT_FILL = make_fill("E8F5E9")   # 新規スロット（内容未生成）
EXISTING_FILL = make_fill("FFFFFF")
TREE_FONT    = Font(size=9)
SINGLE_FONT  = Font(size=9, color="555555")


def load_posts(date_str: str) -> dict:
    """指定日のJSONを読み込む。なければ空dict"""
    json_file = POSTS_DIR / f"{date_str}.json"
    if not json_file.exists():
        return {}
    try:
        return json.loads(json_file.read_text(encoding="utf-8"))
    except Exception:
        return {}


def build_sheet(ws, start_date: datetime, days: int):
    # ── ヘッダー ────────────────────────────────────────
    headers = ["日付", "時刻", "種別", "CTA", "投稿1", "投稿2（ツリー）", "投稿3（ツリー）"]
    col_widths = [12, 7, 6, 5, 60, 60, 60]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── データ行 ────────────────────────────────────────
    row = 2
    for day_offset in range(days):
        date = start_date + timedelta(days=day_offset)
        date_str = date.strftime("%Y-%m-%d")
        date_label = date.strftime("%m/%d(%a)").replace(
            "Mon", "月").replace("Tue", "火").replace("Wed", "水").replace(
            "Thu", "木").replace("Fri", "金").replace("Sat", "土").replace("Sun", "日")

        posts_data = load_posts(date_str)
        color = DATE_COLORS[day_offset % 2]
        date_fill = make_fill(color)

        slots = POST_SCHEDULE  # 時刻順に並んだリスト

        first_row_of_date = row
        for slot in slots:
            info = SLOT_PLAN.get(slot, {"type": "tree", "cta": False})
            existing = posts_data.get(slot)  # リストまたはNone

            kind = "ツリー" if info["type"] == "tree" else "単体"
            cta  = "●" if info["cta"] else ""

            p1 = existing[0] if existing and len(existing) > 0 else ""
            p2 = existing[1] if existing and len(existing) > 1 else ""
            p3 = existing[2] if existing and len(existing) > 2 else ""

            # セル背景
            if info["cta"]:
                row_fill = CTA_FILL
            elif existing:
                row_fill = date_fill
            else:
                row_fill = NEW_SLOT_FILL  # 未生成スロット

            values = [date_label, slot, kind, cta, p1, p2, p3]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = make_border()
                cell.font = Font(size=9)

            ws.row_dimensions[row].height = 60
            row += 1

        # 日付セルを結合
        if row - 1 >= first_row_of_date:
            ws.merge_cells(
                start_row=first_row_of_date, start_column=1,
                end_row=row - 1, end_column=1
            )
            date_cell = ws.cell(row=first_row_of_date, column=1)
            date_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            date_cell.font = Font(bold=True, size=9)

    return row - 2  # 総行数


def main():
    args = sys.argv[1:]

    if args:
        start_date = datetime.strptime(args[0], "%Y-%m-%d")
    else:
        start_date = datetime.now() + timedelta(days=1)
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)

    days = int(args[1]) if len(args) >= 2 else 30

    date_range = f"{start_date.strftime('%Y%m%d')}-{(start_date + timedelta(days=days-1)).strftime('%Y%m%d')}"
    filename = f"コンサル_投稿確認シート_{date_range}.xlsx"
    out_path = DESKTOP / filename

    print(f"\nThreads投稿確認シート 出力中...")
    print(f"  期間: {start_date.strftime('%Y-%m-%d')} ～ {(start_date + timedelta(days=days-1)).strftime('%Y-%m-%d')} ({days}日間)")
    print(f"  スロット数/日: {len(POST_SCHEDULE)} 件")
    print(f"  総行数: {days * len(POST_SCHEDULE)} 行")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投稿スケジュール"

    total = build_sheet(ws, start_date, days)

    # 凡例シート
    ws2 = wb.create_sheet("凡例")
    legends = [
        ("背景色", "意味"),
        ("黄色", "CTA付きスロット（フォロー or LINE登録の文言あり）"),
        ("薄緑", "新規スロット（内容未生成 → 要編集 or generate_day.py で生成）"),
        ("白 / 薄青", "既存スロット（生成済み）"),
    ]
    for r, (a, b) in enumerate(legends, 1):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=(r==1))
        ws2.cell(row=r, column=2, value=b)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 60

    wb.save(out_path)
    print(f"\n[完了] 保存先: {out_path}")
    print(f"  ファイル名: {filename}")


if __name__ == "__main__":
    main()
